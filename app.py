"""
Ultimate Dashboard Ekspor-Impor Indonesia + Neraca Pembayaran SEKI BI
Sumber Data:
1. DB BPS Lokal (ekspor_impor_bps.db) - Ekspor/Impor Nasional & EWS
2. ITC Trade Map                 - Mirroring Asimetri (data_trademap.xlsx)
3. SEKI Bank Indonesia           - Neraca Pembayaran (bop_indonesia.db)

Deploy: Railway / Render
  - Upload ekspor_impor_bps.db, data_trademap.xlsx, dan bop_indonesia.db via Volume atau embed di repo
"""

import os, re, sqlite3
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from dash import Dash, dcc, html, Input, Output, State, dash_table
import plotly.graph_objects as go
import plotly.express as px
from datetime import datetime

# ─────────────────────────────────────────────────────────────────
#  KONFIGURASI UTAMA
# ─────────────────────────────────────────────────────────────────
DATA_DIR     = os.environ.get("DATA_DIR", os.path.dirname(__file__))
BPS_DB_PATH  = os.path.join(DATA_DIR, os.environ.get("BPS_DB_FILE", "ekspor_impor_bps.db"))
BOP_DB_PATH  = os.path.join(DATA_DIR, os.environ.get("BOP_DB_FILE", "bop_indonesia.db"))
TM_XLSX      = os.path.join(DATA_DIR, os.environ.get("TM_XLSX_FILE", "data_trademap.xlsx"))

# NAMA TABEL DI DALAM DB BPS (Silakan sesuaikan jika berbeda)
BPS_TABLE    = "data_eksim" 

PORT         = int(os.environ.get("PORT", 8050))

HS_ALL       = [str(i).zfill(2) for i in range(1, 100)]
TAHUN_SAAT_INI = datetime.now().year
TAHUN_TERSEDIA = list(range(2015, TAHUN_SAAT_INI + 1))

PERIODE_OPSI = [
    {"label": "Tahunan",   "value": "tahunan"},
    {"label": "Januari",   "value": "1"},  {"label": "Februari",  "value": "2"},
    {"label": "Maret",     "value": "3"},  {"label": "April",     "value": "4"},
    {"label": "Mei",       "value": "5"},  {"label": "Juni",      "value": "6"},
    {"label": "Juli",      "value": "7"},  {"label": "Agustus",   "value": "8"},
    {"label": "September", "value": "9"},  {"label": "Oktober",   "value": "10"},
    {"label": "November",  "value": "11"}, {"label": "Desember",  "value": "12"},
]

PARTNER_LIST = [
    "China", "Amerika Serikat", "Jepang", "Singapura",
    "India", "Malaysia", "Korea Selatan", "Australia",
    "Jerman", "Belanda", "Thailand", "Vietnam",
]

HS_DESC = {
    "01":"Binatang Hidup","02":"Daging & Produk Daging","03":"Ikan & Produk Ikan",
    "04":"Produk Susu & Telur","05":"Produk Hewani Lainnya","06":"Tanaman Hidup & Bunga",
    "07":"Sayuran","08":"Buah-buahan","09":"Kopi, Teh & Rempah","10":"Serealia",
    "11":"Produk Penggilingan","12":"Biji Minyak","13":"Getah & Resin",
    "14":"Bahan Nabati Lainnya","15":"Lemak & Minyak Nabati/Hewani","16":"Olahan Daging & Ikan",
    "17":"Gula & Kembang Gula","18":"Kakao & Olahannya","19":"Olahan Sereal & Tepung",
    "20":"Olahan Sayur & Buah","21":"Aneka Olahan Pangan","22":"Minuman & Cuka",
    "23":"Ampas Industri Pangan","24":"Tembakau","25":"Garam, Belerang & Batu",
    "26":"Bijih, Terak & Abu","27":"Bahan Bakar Mineral & Minyak Bumi",
    "28":"Kimia Anorganik","29":"Kimia Organik","30":"Produk Farmasi",
    "31":"Pupuk","32":"Cat, Tinta & Vernis","33":"Minyak Atsiri & Kosmetik",
    "34":"Sabun & Deterjen","35":"Albuminoid & Pati","36":"Bahan Peledak",
    "37":"Produk Foto","38":"Kimia Lainnya","39":"Plastik & Barang Plastik",
    "40":"Karet & Barang Karet","41":"Kulit Mentah & Samak","42":"Barang Kulit & Tas",
    "43":"Bulu Binatang","44":"Kayu & Produk Kayu","45":"Gabus",
    "46":"Produk Anyaman","47":"Bubur Kayu (Pulp)","48":"Kertas & Karton",
    "49":"Buku & Produk Cetak","50":"Sutra","51":"Wol & Bulu Halus",
    "52":"Kapas","53":"Serat Tekstil Lainnya","54":"Filamen Buatan",
    "55":"Serat Staple Buatan","56":"Benang & Kain Khusus","57":"Karpet & Alas Lantai",
    "58":"Kain Tenunan Khusus","59":"Kain Teknik","60":"Kain Rajut",
    "61":"Pakaian Rajut","62":"Pakaian Tenun","63":"Tekstil Rumah Tangga",
    "64":"Alas Kaki","65":"Topi & Aksesori Kepala","66":"Payung & Tongkat",
    "67":"Bulu Olahan & Bunga Buatan","68":"Barang dari Batu & Semen",
    "69":"Produk Keramik","70":"Kaca & Produk Kaca",
    "71":"Batu Permata & Logam Mulia","72":"Besi & Baja",
    "73":"Barang dari Besi/Baja","74":"Tembaga & Produknya",
    "75":"Nikel & Produknya","76":"Aluminium & Produknya",
    "78":"Timbal & Produknya","79":"Seng & Produknya","80":"Timah & Produknya",
    "81":"Logam Dasar Lainnya","82":"Perkakas & Peralatan Logam",
    "83":"Barang Logam Lainnya","84":"Mesin & Perlengkapan Mekanik",
    "85":"Mesin & Perlengkapan Listrik","86":"Kereta Api & Komponen",
    "87":"Kendaraan Bermotor","88":"Pesawat Udara & Komponen",
    "89":"Kapal & Perahu","90":"Instrumen Optik & Medis",
    "91":"Jam & Arloji","92":"Instrumen Musik","93":"Senjata & Amunisi",
    "94":"Furnitur & Perlengkapan","95":"Mainan & Perlengkapan Olahraga",
    "96":"Produk Manufaktur Lainnya","97":"Karya Seni & Antik","99":"Barang Lainnya",
}

# ─────────────────────────────────────────────────────────────────
#  SEKI — KONFIGURASI INDIKATOR
# ─────────────────────────────────────────────────────────────────
BOP_MAIN_ITEMS = {
    1:"Transaksi Berjalan", 2:"Barang", 17:"Jasa",
    20:"Pendapatan Primer", 23:"Pendapatan Sekunder",
    26:"Transaksi Modal", 29:"Transaksi Finansial",
    32:"Investasi Langsung", 35:"Investasi Portofolio",
    40:"Derivatif Finansial", 41:"Investasi Lainnya",
    46:"Total (I+II+III)", 47:"Selisih Perhitungan",
    48:"Neraca Keseluruhan", 54:"Cadangan Devisa",
    56:"CA % PDB",
}
BOP_DD_OPTIONS = [{"label": v, "value": k} for k, v in BOP_MAIN_ITEMS.items()]

# ─────────────────────────────────────────────────────────────────
#  NORMALISASI NEGARA
# ─────────────────────────────────────────────────────────────────
_NEGARA_KW = {
    "China":           ["china", "cina", "tiongkok", "zhongguo"],
    "Amerika Serikat": ["united states", "america", "u.s.a", "u.s.", "serikat", "usa", "u.s.a."],
    "Jepang":          ["japan", "jepang", "nippon"],
    "Singapura":       ["singapore", "singapura"],
    "India":           ["india"],
    "Malaysia":        ["malaysia"],
    "Korea Selatan":   ["korea"],
    "Australia":       ["australia"],
    "Jerman":          ["germany", "jerman", "deutschland"],
    "Belanda":         ["netherlands", "belanda", "holland"],
    "Thailand":        ["thailand"],
    "Vietnam":         ["vietnam", "viet nam", "viet-nam"],
}

def normalize_negara(nama: str) -> str:
    if not nama: return nama
    s   = str(nama).strip()
    low = s.lower()
    for display, kws in _NEGARA_KW.items():
        if low in kws or any(kw in low for kw in kws):
            return display
    return s

def clean_hs(raw) -> str:
    if pd.isna(raw) or str(raw).strip() in ("", "nan"): return ""
    s = str(raw).strip()
    m = re.search(r'\d+', s)
    if m: return m.group(0)[:2].zfill(2)
    try: return str(int(float(s))).zfill(2)
    except (ValueError, TypeError): return s[:2].zfill(2)

def get_periode_params(pilihan):
    return ("2", "") if pilihan == "tahunan" else ("1", pilihan)

# ─────────────────────────────────────────────────────────────────
#  FUNGSI DATABASE BPS (PENGGANTI API)
# ─────────────────────────────────────────────────────────────────
# NAMA TABEL DI DALAM DB BPS
BPS_TABLE = "exim_data" 

def check_bps_db():
    return os.path.exists(BPS_DB_PATH)

def fetch_bps_db(sumber, tahun, tipe, bulan=""):
    """Mengambil seluruh data BPS berdasarkan filter dari SQLite."""
    if not check_bps_db():
        return pd.DataFrame()
    try:
        conn = sqlite3.connect(BPS_DB_PATH)
        # Mapping parameter (1/2) menjadi teks sesuai isi DB (Ekspor/Impor)
        jenis_transaksi = "Ekspor" if str(sumber) == "1" else "Impor"
        
        # Query disesuaikan dengan kolom database Anda: ctr -> negara, netweight -> berat
        query = f"SELECT kode_hs as kodehs, ctr as negara, value, netweight as berat FROM {BPS_TABLE} WHERE jenis_transaksi = ? AND tahun = ?"
        params = [jenis_transaksi, str(tahun)]
        
        # Filter bulan jika user memilih periode bulan tertentu (misal: "09")
        if tipe == "1" and bulan:
            query += " AND bulan_kode = ?"
            params.append(str(bulan).zfill(2)) 
            
        df = pd.read_sql_query(query, conn, params=params)
        conn.close()

        if not df.empty:
            df["kodehs"] = df["kodehs"].apply(clean_hs)
            df["negara"] = df["negara"].apply(normalize_negara)
            df["value"] = pd.to_numeric(df["value"], errors="coerce").fillna(0)
            df["berat"] = pd.to_numeric(df["berat"], errors="coerce").fillna(0)
        return df
    except Exception as e:
        print(f"Error Database BPS: {e}")
        return pd.DataFrame()

def fetch_hist_bps_db(sumber, hs, tipe, bulan=""):
    """Fungsi khusus super-cepat untuk query historis per-HS."""
    if not check_bps_db():
        return pd.DataFrame()
    try:
        conn = sqlite3.connect(BPS_DB_PATH)
        jenis_transaksi = "Ekspor" if str(sumber) == "1" else "Impor"
        
        query = f"SELECT tahun, ctr as negara, value FROM {BPS_TABLE} WHERE jenis_transaksi = ? AND kode_hs = ?"
        params = [jenis_transaksi, str(hs).zfill(2)]
        
        if tipe == "1" and bulan:
            query += " AND bulan_kode = ?"
            params.append(str(bulan).zfill(2))
            
        df = pd.read_sql_query(query, conn, params=params)
        conn.close()

        if not df.empty:
            df["negara"] = df["negara"].apply(normalize_negara)
            df["value"] = pd.to_numeric(df["value"], errors="coerce").fillna(0)
        return df
    except Exception as e:
        print(f"Error Database Historis BPS: {e}")
        return pd.DataFrame()

# ─────────────────────────────────────────────────────────────────
#  SEKI — FUNGSI DATABASE
# ─────────────────────────────────────────────────────────────────
def bop_ok():
    return os.path.exists(BOP_DB_PATH)

def bop_query(sql, params=()):
    try:
        conn = sqlite3.connect(BOP_DB_PATH)
        df   = pd.read_sql_query(sql, conn, params=params)
        conn.close()
        return df
    except Exception:
        return pd.DataFrame()

def bop_years():
    df = bop_query("SELECT DISTINCT year FROM bop_quarterly ORDER BY year")
    return df["year"].tolist() if not df.empty else list(range(2004, 2026))

def bop_latest():
    df = bop_query("""SELECT period FROM bop_quarterly
                      WHERE value_mn_usd IS NOT NULL
                      ORDER BY year DESC, quarter DESC LIMIT 1""")
    return df["period"].iloc[0] if not df.empty else "-"

def bop_series(item_ids, y1, y2, freq):
    ph  = ",".join("?" * len(item_ids))
    sql = f"""SELECT item_id, keterangan, items_en, year, quarter, period, value_mn_usd
              FROM bop_quarterly
              WHERE item_id IN ({ph}) AND year >= ? AND year <= ?
              ORDER BY item_id, year, quarter"""
    df = bop_query(sql, tuple(item_ids) + (y1, y2))
    if df.empty or freq == "quarterly": return df
    
    ratio = {54, 55, 56, 57, 58}
    parts = []
    for iid, grp in df.groupby("item_id"):
        if iid in ratio:
            r = grp[grp["quarter"] == "Q4"][["item_id","keterangan","items_en","year","value_mn_usd"]].copy()
        else:
            r = grp.groupby(["item_id","keterangan","items_en","year"], as_index=False)["value_mn_usd"].sum()
        parts.append(r)
    return pd.concat(parts, ignore_index=True) if parts else pd.DataFrame()

def bop_latest_val(item_id):
    df = bop_query("""SELECT value_mn_usd FROM bop_quarterly
                      WHERE item_id=? AND value_mn_usd IS NOT NULL
                      ORDER BY year DESC, quarter DESC LIMIT 1""", (item_id,))
    return float(df["value_mn_usd"].iloc[0]) if not df.empty else None

_BOP_OK     = bop_ok()
_BOP_YEARS  = bop_years() if _BOP_OK else list(range(2004, 2026))
_BOP_LATEST = bop_latest() if _BOP_OK else "-"

# ─────────────────────────────────────────────────────────────────
#  EWS & KURS
# ─────────────────────────────────────────────────────────────────
def calculate_ews(df):
    if df.empty: return df
    
    # Penanganan aman untuk nilai nol/null
    df["harga"] = df.apply(lambda row: row["value"] / row["berat"] if row.get("berat", 0) > 0 else 0, axis=1)
    
    m_val, s_val = df["value"].mean(), df["value"].std()
    df["z_score"] = 0 if pd.isna(s_val) or s_val == 0 else (df["value"] - m_val) / s_val
    
    df_vh = df[df["harga"] > 0]
    m_h = df_vh["harga"].mean() if not df_vh.empty else 0
    s_h = df_vh["harga"].std()  if not df_vh.empty else 0
    df["z_score_harga"] = df.apply(lambda row: 0 if pd.isna(s_h) or s_h == 0 or row["harga"] == 0 else (row["harga"] - m_h) / s_h, axis=1)
    
    df["status_ews"] = "Normal"
    df.loc[df["z_score"] >  1.5, "status_ews"] = "🔴 Batas Atas Nilai"
    df.loc[df["z_score"] < -0.5, "status_ews"] = "🟡 Batas Bawah Nilai"
    df.loc[df["z_score_harga"] > 2.0, "status_ews"] = "🟣 Anomali Harga (Terlalu Mahal)"
    
    mask = (df["z_score"] > 1.5) & (df["z_score_harga"] > 2.0)
    df.loc[mask, "status_ews"] = "🚨 KRITIS: Nilai Konsentrasi & Harga Spike"
    return df

KURS_HIST = {
    "2015":13389,"2016":13307,"2017":13384,"2018":14236,"2019":14147,
    "2020":14582,"2021":14311,"2022":14850,"2023":15255,"2024":15700,"2025":15850,
}

def get_kurs(tahun, periode):
    # Fallback to static if not currently querying live API
    k   = KURS_HIST.get(str(tahun), 15000)
    lbl = "Tahun" if periode == "tahunan" else "Bulan"
    return k, f"KURS USD/IDR (RATA-RATA {lbl.upper()} {tahun})"

# ─────────────────────────────────────────────────────────────────
#  XLSX EXPORT HELPER — TOP 15 KOMODITAS
# ─────────────────────────────────────────────────────────────────
def build_top15_xlsx(df_top15: pd.DataFrame, jenis: str, tahun: str, unit_lbl: str, is_ekspor: bool) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = f"Top15 {jenis}"

    accent_hex = "1F6E3A" if is_ekspor else "0969DA"
    thin_side  = Side(style="thin", color="D0D7DE")
    border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    # Header & Styling code
    ws.merge_cells("A1:F1")
    ws["A1"] = f"TOP 15 KOMODITAS {jenis.upper()} INDONESIA — TAHUN {tahun}"
    ws["A1"].font      = Font(bold=True, size=13, color="FFFFFF", name="Arial")
    ws["A1"].fill      = PatternFill("solid", start_color=accent_hex)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:F2")
    ws["A2"] = f"Sumber: BPS DB  |  Satuan: {unit_lbl}  |  Diunduh: {datetime.now().strftime('%d %b %Y %H:%M')}"
    ws["A2"].font      = Font(italic=True, size=9, color="6E7781", name="Arial")
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[2].height = 16

    headers = ["No", "Kode HS", "Deskripsi Komoditas", f"Nilai ({unit_lbl})", "Volume (kg)", "Share (%)"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=h)
        cell.font      = Font(bold=True, color="FFFFFF", size=10, name="Arial")
        cell.fill      = PatternFill("solid", start_color="2D333B")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = border_all
    ws.row_dimensions[3].height = 20

    total_val = df_top15["value"].sum()
    for i, row in enumerate(df_top15.itertuples(index=False), 1):
        fill_hex  = "F6F8FA" if i % 2 == 0 else "FFFFFF"
        nilai     = getattr(row, "value", 0)
        volume    = getattr(row, "berat", 0)
        share_pct = (nilai / total_val * 100) if total_val > 0 else 0
        vals = [i, getattr(row, "kodehs", ""), getattr(row, "deskripsi", ""), nilai, volume, share_pct]
        
        for col_idx, v in enumerate(vals, 1):
            cell = ws.cell(row=i + 3, column=col_idx, value=v)
            cell.fill   = PatternFill("solid", start_color=fill_hex)
            cell.border = border_all
            cell.font   = Font(name="Arial", size=10)
            if col_idx == 1: cell.alignment = Alignment(horizontal="center")
            elif col_idx == 3: cell.alignment = Alignment(horizontal="left")
            elif col_idx == 4:
                cell.number_format = "#,##0.00"
                cell.alignment     = Alignment(horizontal="right")
            elif col_idx == 5:
                cell.number_format = "#,##0"
                cell.alignment     = Alignment(horizontal="right")
            elif col_idx == 6:
                cell.number_format = "0.0%"
                cell.value         = share_pct / 100
                cell.alignment     = Alignment(horizontal="right")

    total_row = len(df_top15) + 4
    ws.merge_cells(f"A{total_row}:C{total_row}")
    ws[f"A{total_row}"] = "TOTAL TOP 15"
    ws[f"A{total_row}"].font      = Font(bold=True, size=10, color="FFFFFF", name="Arial")
    ws[f"A{total_row}"].fill      = PatternFill("solid", start_color=accent_hex)
    ws[f"A{total_row}"].alignment = Alignment(horizontal="center")

    ws[f"D{total_row}"] = total_val
    ws[f"D{total_row}"].number_format = "#,##0.00"
    ws[f"D{total_row}"].font = Font(bold=True, color="FFFFFF", name="Arial")
    ws[f"D{total_row}"].fill = PatternFill("solid", start_color=accent_hex)
    ws[f"D{total_row}"].alignment = Alignment(horizontal="right")

    ws[f"E{total_row}"] = df_top15["berat"].sum()
    ws[f"E{total_row}"].number_format = "#,##0"
    ws[f"E{total_row}"].font = Font(bold=True, color="FFFFFF", name="Arial")
    ws[f"E{total_row}"].fill = PatternFill("solid", start_color=accent_hex)
    ws[f"E{total_row}"].alignment = Alignment(horizontal="right")

    ws[f"F{total_row}"] = 1.0
    ws[f"F{total_row}"].number_format = "0.0%"
    ws[f"F{total_row}"].font = Font(bold=True, color="FFFFFF", name="Arial")
    ws[f"F{total_row}"].fill = PatternFill("solid", start_color=accent_hex)
    ws[f"F{total_row}"].alignment = Alignment(horizontal="right")

    for col_idx in range(1, 7):
        ws.cell(row=total_row, column=col_idx).border = border_all

    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 38
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 12

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()

# ─────────────────────────────────────────────────────────────────
#  TEMA & CHART HELPERS
# ─────────────────────────────────────────────────────────────────
THEMES = {
    "dark": {
        "bg":"#0d1117","card":"#161b22","border":"#30363d","green":"#3fb950",
        "red":"#f78166","blue":"#79c0ff","accent":"#58a6ff","yellow":"#e3b341",
        "text":"#e6edf3","muted":"#8b949e","grid":"#21262d","purple":"#bc8cff",
        "teal":"#39d0d8","orange":"#ffa657",
    },
    "light": {
        "bg":"#f6f8fa","card":"#ffffff","border":"#d0d7de","green":"#1a7f37",
        "red":"#cf222e","blue":"#0969da","accent":"#0969da","yellow":"#bf8700",
        "text":"#24292f","muted":"#57606a","grid":"#e1e4e8","purple":"#8250df",
        "teal":"#0e7490","orange":"#bc4c00",
    },
}
FONT = "'IBM Plex Mono','Courier New',monospace"

def base_chart(t):
    return dict(paper_bgcolor=t["card"], plot_bgcolor=t["card"],
                font=dict(color=t["text"], family=FONT, size=11),
                margin=dict(l=10, r=20, t=30, b=30))

def empty_fig(t, msg="Klik MUAT DATA untuk memulai"):
    fig = go.Figure()
    fig.add_annotation(text=msg, x=.5, y=.5, xref="paper", yref="paper",
                       showarrow=False, font=dict(color=t["muted"], size=13, family=FONT))
    fig.update_layout(**base_chart(t), height=340)
    return fig

# ─────────────────────────────────────────────────────────────────
#  LAYOUT HELPERS
# ─────────────────────────────────────────────────────────────────
_dd = {"color": "#000"}

def _btn_download(btn_id, label, color="#1f6e3a"):
    return html.Button(label, id=btn_id,
                       style={"padding":"6px 14px","cursor":"pointer","fontWeight":"bold",
                              "backgroundColor":color,"color":"#ffffff",
                              "border":"none","borderRadius":"4px",
                              "fontFamily":FONT,"fontSize":"11px"})

def _card_header_with_downloads(title, btn_csv_id, btn_xlsx_id, dl_csv_id, dl_xlsx_id,
                                 csv_color="#1f6e3a", xlsx_color="#0969da"):
    return html.Div(
        style={"display":"flex","justifyContent":"space-between",
               "alignItems":"center","marginBottom":"12px"},
        children=[
            html.Div(title, style={"fontSize":"10px","fontWeight":"bold","letterSpacing":"1px"}),
            html.Div(style={"display":"flex","gap":"8px","alignItems":"center"}, children=[
                _btn_download(btn_csv_id,  "⬇ CSV",  csv_color),
                _btn_download(btn_xlsx_id, "⬇ XLSX", xlsx_color),
                dcc.Download(id=dl_csv_id),
                dcc.Download(id=dl_xlsx_id),
            ]),
        ],
    )

# ─────────────────────────────────────────────────────────────────
#  LAYOUT
# ─────────────────────────────────────────────────────────────────
app = Dash(__name__)
server = app.server

app.title = "Trade Intelligence | BPS · Trade Map · SEKI BI"

app.layout = html.Div(id="main-container", children=[

    # ── Stores ──────────────────────────────────────────────────
    dcc.Store(id="theme-store", data="dark"),
    dcc.Store(id="store-bps"),
    dcc.Store(id="store-top15"),

    # ── Header ──────────────────────────────────────────────────
    html.Div(id="hdr", style={"display":"flex","justifyContent":"space-between",
                               "alignItems":"center","marginBottom":"20px"}, children=[
        html.Div([
            html.Span("NEXOS | ", style={"fontWeight":"bold","fontSize":"18px"}),
            html.Span("BPS · ITC Trade Map · SEKI Bank Indonesia", style={"fontSize":"16px"}),
        ]),
        html.Div(style={"display":"flex","gap":"15px","alignItems":"center"}, children=[
            html.Div(id="ts", style={"fontSize":"12px"}),
            html.Button("🌓 Tema", id="btn-theme",
                        style={"padding":"8px 16px","cursor":"pointer","fontFamily":FONT,
                               "borderRadius":"4px","backgroundColor":"transparent",
                               "color":"inherit","fontWeight":"bold"}),
        ]),
    ]),

    # ── Filter BPS ──────────────────────────────────────────────
    html.Div(id="filter-card", children=[
        html.Div(style={"display": "flex", "justifyContent": "space-between"}, children=[
            html.Div("1. KONTROL DATA BPS (LOCAL DATABASE)",
                     style={"fontSize":"10px","letterSpacing":"2px",
                            "marginBottom":"12px","fontWeight":"bold"}),
            html.Div(id="db-badge", style={"fontSize":"11px", "fontWeight":"bold"})
        ]),
        html.Div(style={"display":"grid",
                         "gridTemplateColumns":"repeat(auto-fit,minmax(150px,1fr))",
                         "gap":"14px","alignItems":"end"}, children=[
            html.Div([html.Label("Tahun", style={"fontSize":"11px"}),
                      dcc.Dropdown(id="dd-tahun",
                                   options=[{"label":str(t),"value":str(t)} for t in reversed(TAHUN_TERSEDIA)],
                                   value=str(TAHUN_TERSEDIA[-2]),
                                   clearable=False, style=_dd)]),
            html.Div([html.Label("Periode", style={"fontSize":"11px"}),
                      dcc.Dropdown(id="dd-periode", options=PERIODE_OPSI,
                                   value="tahunan", clearable=False, style=_dd)]),
            html.Div([html.Label("Jenis Perdagangan", style={"fontSize":"11px"}),
                      dcc.RadioItems(id="radio-sumber",
                                     options=[{"label":" Ekspor","value":"1"},
                                              {"label":" Impor","value":"2"}],
                                     value="1", inline=True,
                                     labelStyle={"marginRight":"15px","cursor":"pointer",
                                                 "color":"inherit","fontWeight":"bold"})]),
            html.Div([html.Label("Filter Negara", style={"fontSize":"11px"}),
                      dcc.Dropdown(id="input-negara", placeholder="Semua Negara...", style=_dd)]),
            html.Div([html.Label("Filter HS", style={"fontSize":"11px"}),
                      dcc.Dropdown(id="input-hs",
                                   options=[{"label":f"HS {h}","value":h} for h in HS_ALL],
                                   placeholder="Semua HS...", clearable=True, style=_dd)]),
            html.Div([html.Label("Satuan", style={"fontSize":"11px"}),
                      dcc.RadioItems(id="radio-unit",
                                     options=[{"label":" USD","value":1},
                                              {"label":" Miliar USD","value":1e9}],
                                     value=1, inline=True,
                                     labelStyle={"marginRight":"15px","cursor":"pointer",
                                                 "color":"inherit","fontWeight":"bold"})]),
            html.Button("▶ MUAT BPS", id="btn-load-bps",
                        style={"padding":"10px","cursor":"pointer","fontWeight":"bold",
                               "backgroundColor":"#1f6e3a","color":"#ffffff",
                               "border":"none","borderRadius":"4px"}),
        ]),
    ]),

    html.Div(id="status-bps", style={"margin":"12px 0","fontSize":"12px","minHeight":"16px"}),

    # ── Tabs ────────────────────────────────────────────────────
    dcc.Tabs(id="tabs", value="tab-ringkasan", children=[
        
        # ── TAB 1 — Ringkasan BPS ───────────────────────────────
        dcc.Tab(label="📊 Ringkasan BPS", value="tab-ringkasan", id="tab-1", children=[
            html.Div(style={"paddingTop":"20px"}, children=[
                html.Div(id="kpi", style={"display":"grid",
                                          "gridTemplateColumns":"repeat(auto-fit,minmax(200px,1fr))",
                                          "gap":"14px","marginBottom":"18px"}),
                html.Div(style={"display":"grid","gridTemplateColumns":"1fr 1fr",
                                 "gap":"16px","marginBottom":"16px"}, children=[
                    html.Div(id="hist-card", children=[
                        html.Div(f"HISTORIS TREN ({TAHUN_TERSEDIA[0]}–{TAHUN_SAAT_INI})",
                                 style={"fontSize":"12px","fontWeight":"bold","marginBottom":"12px"}),
                        html.Div(style={"display":"grid",
                                         "gridTemplateColumns":"repeat(auto-fit,minmax(130px,1fr))",
                                         "gap":"10px","marginBottom":"15px"}, children=[
                            dcc.Dropdown(id="hist-hs",
                                         options=[{"label":f"HS {h}","value":h} for h in HS_ALL],
                                         placeholder="Pilih HS spesifik...", style=_dd),
                            dcc.Dropdown(id="hist-negara", placeholder="Semua Negara...", style=_dd),
                            dcc.RadioItems(id="hist-metric",
                                           options=[{"label":" Nilai","value":"nilai"},
                                                    {"label":" YoY %","value":"yoy"}],
                                           value="nilai", inline=True,
                                           labelStyle={"marginRight":"10px", "color":"inherit"}),
                            html.Button("Tampilkan Histori", id="btn-hist",
                                        style={"padding":"8px","backgroundColor":"#e3b341",
                                               "border":"none","borderRadius":"4px",
                                               "fontWeight":"bold","cursor":"pointer"}),
                        ]),
                        dcc.Loading(dcc.Graph(id="g-hist", config={"displayModeBar":False}, style={"height":"350px"})),
                    ]),
                    html.Div(id="chart-card-1", children=[
                        _card_header_with_downloads(
                            title        = "TOP 15 KOMODITAS (HS)",
                            btn_csv_id   = "btn-dl-top15-csv",
                            btn_xlsx_id  = "btn-dl-top15-xlsx",
                            dl_csv_id    = "dl-top15-csv",
                            dl_xlsx_id   = "dl-top15-xlsx",
                        ),
                        dcc.Graph(id="g-kmd", config={"displayModeBar":False}),
                    ]),
                ]),
                html.Div(style={"display":"grid","gridTemplateColumns":"1fr 1fr", "gap":"16px"}, children=[
                    html.Div(id="chart-card-2", children=[
                        html.Div("TOP NEGARA MITRA", style={"fontSize":"10px","fontWeight":"bold","marginBottom":"8px"}),
                        dcc.Graph(id="g-neg", config={"displayModeBar":False}),
                    ]),
                    html.Div(id="chart-card-3", children=[
                        html.Div("SHARE KOMODITAS", style={"fontSize":"10px","fontWeight":"bold","marginBottom":"8px"}),
                        dcc.Graph(id="g-pie", config={"displayModeBar":False}),
                    ]),
                ]),
            ]),
        ]),

        # ── TAB 2 — Data Lengkap ────────────────────────────────
        dcc.Tab(label="🗄️ Data Lengkap BPS", value="tab-tabel", id="tab-2", children=[
            html.Div(id="table-card-full", style={"paddingTop":"20px"}, children=[
                html.Div(style={"display":"flex","justifyContent":"flex-end", "marginBottom":"10px"}, children=[
                    html.Button("⬇ Download CSV", id="btn-dl-bps",
                                style={"padding":"8px 16px","cursor":"pointer",
                                       "fontWeight":"bold","backgroundColor":"#1f6e3a",
                                       "color":"#ffffff","border":"none",
                                       "borderRadius":"4px","fontFamily":FONT}),
                    dcc.Download(id="dl-bps"),
                ]),
                dash_table.DataTable(id="tabel-full", page_size=20, sort_action="native", filter_action="native"),
            ]),
        ]),

        # ── TAB 3 — EWS ─────────────────────────────────────────
        dcc.Tab(label="⚠️ Early Warning System", value="tab-ews", id="tab-3", children=[
            html.Div(id="ews-card", style={"paddingTop":"20px"}, children=[
                html.Div("Deteksi anomali dengan Z-Score. Batas Atas = konsentrasi berlebih. Batas Bawah = underperforming.",
                         style={"fontSize":"11px","marginBottom":"15px", "fontStyle":"italic"}),
                dash_table.DataTable(id="tabel-ews", page_size=15, sort_action="native", filter_action="native"),
            ]),
        ]),

        # ── TAB 4 — Mirroring ───────────────────────────────────
        dcc.Tab(label="🪞 Mirroring (BPS vs Trade Map)", value="tab-mirror", id="tab-4", children=[
            html.Div(id="mirror-container", style={"paddingTop":"20px"}, children=[
                html.Div(style={"background":"rgba(9,105,218,0.1)", "border":"1px solid #0969da",
                                 "padding":"15px","borderRadius":"8px", "marginBottom":"20px"}, children=[
                    html.Div("2. ANALISIS ASIMETRI PENCATATAN",
                             style={"fontSize":"11px","fontWeight":"bold", "marginBottom":"10px","color":"#0969da"}),
                    html.Div(style={"display":"grid", "gridTemplateColumns":"repeat(auto-fit,minmax(200px,1fr))",
                                     "gap":"14px","alignItems":"end"}, children=[
                        html.Div([html.Label("Mitra Dagang", style={"fontSize":"11px"}),
                                  dcc.Dropdown(id="dd-mitra", options=[{"label":k,"value":k} for k in PARTNER_LIST],
                                               placeholder="Pilih Negara...", style=_dd)]),
                        html.Div([html.Label("Satuan Mirroring", style={"fontSize":"11px"}),
                                  dcc.RadioItems(id="radio-unit-mirror",
                                                 options=[{"label":" USD","value":1}, {"label":" Juta USD","value":1e6}],
                                                 value=1e6, inline=True,
                                                 labelStyle={"marginRight":"15px","cursor":"pointer",
                                                             "color":"inherit","fontWeight":"bold"})]),
                        html.Button("▶ JALANKAN MIRRORING", id="btn-mirror",
                                    style={"padding":"10px","cursor":"pointer","fontWeight":"bold",
                                           "backgroundColor":"#0969da","color":"#ffffff",
                                           "border":"none","borderRadius":"4px"}),
                    ]),
                ]),
                html.Div(id="status-mirror", style={"marginBottom":"16px","fontSize":"12px", "fontWeight":"bold"}),
                html.Div(id="kpi-mirror", style={"display":"grid",
                                                 "gridTemplateColumns":"repeat(auto-fit,minmax(280px,1fr))",
                                                 "gap":"14px","marginBottom":"18px"}),
                html.Div(style={"display":"grid","gridTemplateColumns":"1fr 1fr", "gap":"16px","marginBottom":"16px"}, children=[
                    html.Div(id="chart-mirror-1", children=[
                        html.Div("KOMPARASI BPS VS ITC TRADE MAP (10 HS TERBESAR BPS)",
                                 style={"fontSize":"10px","marginBottom":"12px", "fontWeight":"bold"}),
                        dcc.Graph(id="g-compare", config={"displayModeBar":False}),
                    ]),
                    html.Div(id="chart-mirror-2", children=[
                        html.Div("5 HS ASIMETRI TERBESAR", style={"fontSize":"10px","marginBottom":"12px", "fontWeight":"bold"}),
                        dcc.Graph(id="g-discrepancy", config={"displayModeBar":False}),
                    ]),
                ]),
                html.Div(id="table-card-mirror", children=[
                    html.Div(style={"display":"flex","justifyContent":"flex-end", "marginBottom":"10px"}, children=[
                        html.Button("⬇ Download CSV", id="btn-dl-mirror",
                                    style={"padding":"8px 16px","cursor":"pointer","fontWeight":"bold",
                                           "backgroundColor":"#0969da","color":"#ffffff",
                                           "border":"none","borderRadius":"4px","fontFamily":FONT}),
                        dcc.Download(id="dl-mirror"),
                    ]),
                    dash_table.DataTable(id="tabel-mirror", page_size=20, sort_action="native", filter_action="native"),
                ]),
            ]),
        ]),

        # ── TAB 5 — SEKI BI ─────────────────────────────────────
        dcc.Tab(label="🏦 Neraca Pembayaran (SEKI BI)", value="tab-seki", id="tab-5", children=[
            html.Div(style={"paddingTop":"20px"}, children=[
                html.Div(id="seki-db-badge", style={"marginBottom":"14px"}),
                html.Div(id="seki-filter-card", children=[
                    html.Div("3. NERACA PEMBAYARAN — SEKI BANK INDONESIA (2004–2025)",
                             style={"fontSize":"10px","letterSpacing":"2px", "marginBottom":"14px","fontWeight":"bold"}),
                    html.Div(style={"display":"grid", "gridTemplateColumns":"repeat(auto-fit,minmax(160px,1fr))",
                                     "gap":"14px","alignItems":"end"}, children=[
                        html.Div([html.Label("Tahun Awal", style={"fontSize":"11px"}),
                                  dcc.Dropdown(id="seki-y1", options=[{"label":str(y),"value":y} for y in _BOP_YEARS],
                                               value=2015, clearable=False, style=_dd)]),
                        html.Div([html.Label("Tahun Akhir", style={"fontSize":"11px"}),
                                  dcc.Dropdown(id="seki-y2", options=[{"label":str(y),"value":y} for y in _BOP_YEARS],
                                               value=_BOP_YEARS[-1], clearable=False, style=_dd)]),
                        html.Div([html.Label("Frekuensi", style={"fontSize":"11px"}),
                                  dcc.RadioItems(id="seki-freq",
                                                 options=[{"label":" Kuartalan", "value":"quarterly"},
                                                          {"label":" Tahunan", "value":"annual"}],
                                                 value="quarterly", inline=True,
                                                 labelStyle={"marginRight":"12px","color":"inherit","fontWeight":"bold"})]),
                        html.Div([html.Label("Satuan", style={"fontSize":"11px"}),
                                  dcc.RadioItems(id="seki-unit",
                                                 options=[{"label":" Juta USD","value":1}, {"label":" Miliar USD","value":1000}],
                                                 value=1, inline=True,
                                                 labelStyle={"marginRight":"12px","color":"inherit","fontWeight":"bold"})]),
                        html.Button("▶ TAMPILKAN", id="btn-seki",
                                    style={"padding":"10px","cursor":"pointer","fontWeight":"bold",
                                           "backgroundColor":"#6e40c9","color":"#ffffff",
                                           "border":"none","borderRadius":"4px"}),
                    ]),
                ]),
                html.Div(id="seki-status", style={"margin":"12px 0","fontSize":"12px","minHeight":"16px"}),
                html.Div(id="seki-kpi", style={"display":"grid", "gridTemplateColumns":"repeat(auto-fit,minmax(210px,1fr))",
                                               "gap":"14px","marginBottom":"20px"}),
                html.Div(style={"display":"grid","gridTemplateColumns":"3fr 2fr", "gap":"16px","marginBottom":"16px"}, children=[
                    html.Div(id="seki-c1", children=[
                        html.Div("TREN TRANSAKSI BERJALAN & KOMPONEN", style={"fontSize":"10px","fontWeight":"bold","marginBottom":"8px"}),
                        dcc.Graph(id="seki-g-ca", config={"displayModeBar":False}, style={"height":"340px"}),
                    ]),
                    html.Div(id="seki-c2", children=[
                        html.Div("DEKOMPOSISI NERACA (TOTAL PERIODE)", style={"fontSize":"10px","fontWeight":"bold","marginBottom":"8px"}),
                        dcc.Graph(id="seki-g-wf", config={"displayModeBar":False}, style={"height":"340px"}),
                    ]),
                ]),
                html.Div(style={"display":"grid","gridTemplateColumns":"1fr 1fr", "gap":"16px","marginBottom":"16px"}, children=[
                    html.Div(id="seki-c3", children=[
                        html.Div("TRANSAKSI FINANSIAL: KOMPONEN INVESTASI", style={"fontSize":"10px","fontWeight":"bold","marginBottom":"8px"}),
                        dcc.Graph(id="seki-g-inv", config={"displayModeBar":False}, style={"height":"320px"}),
                    ]),
                    html.Div(id="seki-c4", children=[
                        html.Div("CADANGAN DEVISA & NERACA KESELURUHAN", style={"fontSize":"10px","fontWeight":"bold","marginBottom":"8px"}),
                        dcc.Graph(id="seki-g-cad", config={"displayModeBar":False}, style={"height":"320px"}),
                    ]),
                ]),
                html.Div(style={"display":"grid","gridTemplateColumns":"1fr 1fr", "gap":"16px","marginBottom":"16px"}, children=[
                    html.Div(id="seki-c5", children=[
                        html.Div(style={"display":"flex","gap":"10px","alignItems":"center", "marginBottom":"10px","flexWrap":"wrap"}, children=[
                            html.Span("KOMPARASI INDIKATOR:", style={"fontSize":"10px","fontWeight":"bold"}),
                            dcc.Dropdown(id="seki-dd-cmp", options=BOP_DD_OPTIONS, value=[1, 29, 48],
                                         multi=True, style={**_dd, "minWidth":"280px"}),
                        ]),
                        dcc.Graph(id="seki-g-cmp", config={"displayModeBar":False}, style={"height":"320px"}),
                    ]),
                    html.Div(id="seki-c6", children=[
                        html.Div("CURRENT ACCOUNT % PDB & CADDEV (BULAN IMPOR)", style={"fontSize":"10px","fontWeight":"bold","marginBottom":"8px"}),
                        dcc.Graph(id="seki-g-pct", config={"displayModeBar":False}, style={"height":"320px"}),
                    ]),
                ]),
                html.Div(id="seki-tbl-card", children=[
                    html.Div(style={"display":"flex","justifyContent":"space-between", "alignItems":"center","marginBottom":"10px"}, children=[
                        html.Span("DATA NERACA PEMBAYARAN LENGKAP", style={"fontSize":"10px","fontWeight":"bold"}),
                        html.Div(style={"display":"flex","gap":"10px"}, children=[
                            dcc.Dropdown(id="seki-tbl-filter", options=BOP_DD_OPTIONS, value=None, multi=True,
                                         placeholder="Filter indikator...", style={**_dd, "minWidth":"240px"}),
                            html.Button("⬇ Download CSV", id="btn-dl-seki",
                                        style={"padding":"8px 14px","cursor":"pointer","fontWeight":"bold",
                                               "backgroundColor":"#6e40c9","color":"#ffffff",
                                               "border":"none","borderRadius":"4px","fontFamily":FONT}),
                            dcc.Download(id="dl-seki"),
                        ]),
                    ]),
                    dash_table.DataTable(id="seki-tbl", page_size=20, sort_action="native", filter_action="native"),
                ]),
            ]),
        ]),
    ]),
    dcc.Interval(id="iv", interval=30_000, n_intervals=0),
])

# ─────────────────────────────────────────────────────────────────
#  CALLBACKS — TEMA
# ─────────────────────────────────────────────────────────────────
@app.callback(Output("theme-store","data"), Input("btn-theme","n_clicks"), State("theme-store","data"))
def toggle_theme(n, cur):
    return "light" if cur == "dark" else "dark"

@app.callback(
    [Output("main-container","style"), Output("filter-card","style"),
     Output("chart-card-1","style"), Output("chart-card-2","style"), Output("chart-card-3","style"), 
     Output("hist-card","style"), Output("chart-mirror-1","style"), Output("chart-mirror-2","style"),
     Output("table-card-mirror","style"), Output("seki-filter-card","style"),
     Output("seki-c1","style"), Output("seki-c2","style"), Output("seki-c3","style"), 
     Output("seki-c4","style"), Output("seki-c5","style"), Output("seki-c6","style"),
     Output("seki-tbl-card","style"), Output("status-bps","style"), Output("status-mirror","style"),
     Output("seki-status","style"), Output("ts","style"),
     Output("tab-1","style"), Output("tab-1","selected_style"),
     Output("tab-2","style"), Output("tab-2","selected_style"),
     Output("tab-3","style"), Output("tab-3","selected_style"),
     Output("tab-4","style"), Output("tab-4","selected_style"),
     Output("tab-5","style"), Output("tab-5","selected_style"),
     Output("db-badge", "children"), Output("db-badge", "style")],
    Input("theme-store","data"),
)
def apply_theme(tn):
    t = THEMES[tn]
    main  = {"background":t["bg"],"minHeight":"100vh","fontFamily":FONT, "color":t["text"],"padding":"24px"}
    card  = {"background":t["card"],"border":f"1px solid {t['border']}", "borderRadius":"8px","padding":"20px"}
    tbase = {"backgroundColor":t["bg"],"color":t["muted"], "borderBottom":f"1px solid {t['border']}","padding":"12px"}
    tsel  = {"backgroundColor":t["card"],"color":t["accent"], "borderTop":f"3px solid {t['accent']}","padding":"12px","fontWeight":"bold"}
    t5sel = {**tsel, "color":t["purple"],"borderTop":f"3px solid {t['purple']}"}
    sm    = {"color":t["muted"]}
    
    db_status = "✅ DB BPS Tersambung" if check_bps_db() else "❌ DB BPS Tidak Ditemukan"
    db_color = t["green"] if check_bps_db() else t["red"]
    
    return (main, card, card, card, card, card, card, card, card, card, card, card, card, card, card, card, card,
            sm, {"color":t["text"]}, {"color":t["purple"]}, sm,
            tbase, tsel, tbase, tsel, tbase, tsel, tbase, tsel, tbase, t5sel,
            db_status, {"color": db_color})

@app.callback(Output("ts","children"), Input("iv","n_intervals"))
def tick(_):
    return datetime.now().strftime("⏱ %d %b %Y  %H:%M")

# ─────────────────────────────────────────────────────────────────
#  CALLBACKS — FETCH BPS
# ─────────────────────────────────────────────────────────────────
@app.callback(
    Output("store-bps","data"), Output("status-bps","children"),
    Input("btn-load-bps","n_clicks"),
    State("dd-tahun","value"), State("dd-periode","value"), State("radio-sumber","value"),
    prevent_initial_call=True,
)
def cb_fetch_bps(_, tahun, periode, sumber):
    tipe, bulan = get_periode_params(periode)
    jenis = "Ekspor" if str(sumber) == "1" else "Impor"
    try:
        df = fetch_bps_db(sumber, tahun, tipe, bulan)
        if df.empty:
            return {}, f"⚠️ Tidak ada data BPS {jenis} {tahun} di Database Lokal."
        return (
            {"data": df.to_dict("records"), "sumber": sumber, "tahun": tahun, "periode": periode},
            f"✅ Data BPS {jenis} {tahun} berhasil ditarik dari Database — {len(df):,} baris.",
        )
    except Exception as e:
        return {}, f"❌ Error BPS DB: {e}"

# ─────────────────────────────────────────────────────────────────
#  CALLBACKS — DASHBOARD BPS (termasuk store-top15)
# ─────────────────────────────────────────────────────────────────
@app.callback(
    [Output("store-top15","data"), Output("kpi","children"),
     Output("g-kmd","figure"), Output("g-neg","figure"), Output("g-pie","figure"),
     Output("tabel-full","data"), Output("tabel-full","columns"), Output("tabel-full","style_data_conditional"),
     Output("tabel-full","style_header"), Output("tabel-full","style_cell"),
     Output("tabel-ews","data"), Output("tabel-ews","columns"), Output("tabel-ews","style_data_conditional"),
     Output("tabel-ews","style_header"), Output("tabel-ews","style_cell"),
     Output("input-negara","options"), Output("hist-negara","options")],
    [Input("store-bps","data"), Input("input-negara","value"), Input("input-hs","value"), Input("radio-unit","value"),
     Input("theme-store","data")],
)
def update_bps(raw, neg_f, hs_f, unit, tn):
    t   = THEMES[tn]
    bc  = base_chart(t)
    hdr = {"backgroundColor":t["card"],"color":t["text"], "fontWeight":"bold","border":f"1px solid {t['border']}"}
    cel = {"backgroundColor":t["bg"],"color":t["text"], "border":f"1px solid {t['border']}","fontFamily":FONT}
    cnd = [{"if":{"row_index":"odd"},"backgroundColor":t["card"]}]

    empty = (None, [], empty_fig(t), empty_fig(t), empty_fig(t),
             [], [], cnd, hdr, cel, [], [], cnd, hdr, cel, [], [])
    if not raw or not raw.get("data"): return empty

    df = pd.DataFrame(raw["data"])
    neg_opts = [{"label":n,"value":n} for n in sorted(df["negara"].unique())]
    if neg_f: df = df[df["negara"] == neg_f]
    if hs_f:  df = df[df["kodehs"] == hs_f]
    df["value"] = df["value"] / unit
    lbl = "Miliar USD" if unit > 1 else "USD"
    fmt = "$,.2f" if unit > 1 else "$,.0f"

    is_ekspor = str(raw.get("sumber")) == "1"
    jenis     = "Ekspor" if is_ekspor else "Impor"
    col       = t["green"] if is_ekspor else t["red"]

    kmd  = df.groupby("kodehs", as_index=False)[["value","berat"]].sum().sort_values("value", ascending=False)
    neg  = df.groupby("negara",  as_index=False)["value"].sum().sort_values("value", ascending=False)
    full = df.groupby(["negara","kodehs"], as_index=False)[["value","berat"]].sum()
    ews  = calculate_ews(kmd.copy())
    full["deskripsi"] = full["kodehs"].map(HS_DESC).fillna("Lainnya")

    top15 = kmd.head(15).copy()
    top15["deskripsi"] = top15["kodehs"].map(HS_DESC).fillna("Lainnya")
    top15_store = {
        "data":       top15.to_dict("records"),
        "unit_label": lbl,
        "tahun":      raw.get("tahun", str(TAHUN_SAAT_INI)),
        "jenis":      jenis,
        "is_ekspor":  is_ekspor,
    }

    kurs, klbl = get_kurs(raw.get("tahun", str(TAHUN_SAAT_INI)), raw.get("periode","tahunan"))
    total = df["value"].sum()

    kpis = [
        html.Div(style={"background":t["card"],"border":f"1px solid {t['border']}",
                         "borderTop":f"3px solid {col}","padding":"16px","borderRadius":"8px"},
                 children=[html.Div(f"TOTAL {jenis.upper()} BPS", style={"fontSize":"10px","color":t["muted"]}),
                           html.Div(f"{total:,.2f} {lbl}", style={"fontSize":"20px","fontWeight":"bold","color":col})]),
        html.Div(style={"background":t["card"],"border":f"1px solid {t['border']}",
                         "borderTop":f"3px solid {t['accent']}","padding":"16px","borderRadius":"8px"},
                 children=[html.Div("KOMODITAS TERBESAR", style={"fontSize":"10px","color":t["muted"]}),
                           html.Div(kmd.iloc[0]["kodehs"] if not kmd.empty else "-",
                                    style={"fontSize":"20px","fontWeight":"bold", "color":t["accent"]})]),
        html.Div(style={"background":t["card"],"border":f"1px solid {t['border']}",
                         "borderTop":f"3px solid {t['yellow']}","padding":"16px","borderRadius":"8px"},
                 children=[html.Div(klbl,style={"fontSize":"10px","color":t["muted"]}),
                           html.Div(f"Rp {kurs:,.0f}" if kurs else "N/A",
                                    style={"fontSize":"20px","fontWeight":"bold", "color":t["yellow"]})]),
    ]

    fk = go.Figure(go.Bar(y=kmd["kodehs"].head(15), x=kmd["value"].head(15),
                          orientation="h", marker_color=col)
                   ).update_layout(**bc, height=350, yaxis=dict(autorange="reversed"))

    fn = go.Figure(go.Bar(y=neg["negara"].head(15), x=neg["value"].head(15),
                          orientation="h", marker_color=t["blue"])
                   ).update_layout(**bc, height=350, yaxis=dict(autorange="reversed"))

    fp = go.Figure(go.Pie(labels=kmd["kodehs"].head(8), values=kmd["value"].head(8),
                          hole=.45, marker=dict(line=dict(color=t["bg"], width=2)))
                   ).update_layout(**bc, height=350, legend=dict(bgcolor="rgba(0,0,0,0)"))

    cols_full = [
        {"name":"Negara Mitra","id":"negara"}, {"name":"HS","id":"kodehs"}, {"name":"Deskripsi","id":"deskripsi"},
        {"name":f"Nilai BPS ({lbl})","id":"value", "type":"numeric","format":{"specifier":fmt}},
    ]
    cols_ews = [
        {"name":"HS","id":"kodehs"},
        {"name":f"Nilai ({lbl})","id":"value", "type":"numeric","format":{"specifier":fmt}},
        {"name":"Volume (kg/ton)*","id":"berat", "type":"numeric","format":{"specifier":",.0f"}},
        {"name":f"Harga Est. ({lbl}/unit)*","id":"harga", "type":"numeric","format":{"specifier":",.4f"}},
        {"name":"Z-Score Nilai","id":"z_score", "type":"numeric","format":{"specifier":".2f"}},
        {"name":"Z-Score Harga","id":"z_score_harga", "type":"numeric","format":{"specifier":".2f"}},
        {"name":"Indikator","id":"status_ews"},
    ]
    cnd_ews = cnd + [
        {"if":{"filter_query":"{status_ews} contains 'Atas'"},  "color":t["red"]},
        {"if":{"filter_query":"{status_ews} contains 'Bawah'"},"color":t["yellow"]},
        {"if":{"filter_query":"{status_ews} contains 'Anomali Harga'"},"color":t["accent"]},
        {"if":{"filter_query":"{status_ews} contains 'KRITIS'"}, "backgroundColor":t["red"],"color":"white"},
    ]
    return (top15_store, kpis, fk, fn, fp, full.to_dict("records"), cols_full, cnd, hdr, cel,
            ews.to_dict("records"), cols_ews, cnd_ews, hdr, cel, neg_opts, neg_opts)

# ─────────────────────────────────────────────────────────────────
#  CALLBACKS — HISTORIS
# ─────────────────────────────────────────────────────────────────
@app.callback(
    Output("g-hist","figure"), Input("btn-hist","n_clicks"),
    State("hist-hs","value"), State("hist-negara","value"),
    State("hist-metric","value"), State("dd-periode","value"),
    State("radio-sumber","value"), State("radio-unit","value"),
    State("theme-store","data"), prevent_initial_call=True,
)
def update_hist(_, hs, negara, metric, periode, sumber, unit, tn):
    t = THEMES[tn]
    if not hs: return empty_fig(t, "⚠️ Pilih HS terlebih dahulu lalu tekan Tampilkan Histori.")
    
    tipe, bulan = get_periode_params(periode)
    df_raw = fetch_hist_bps_db(sumber, hs, tipe, bulan)
    
    if df_raw.empty:
        return empty_fig(t, "⚠️ Tidak ada data historis di Database Lokal.")

    if negara:
        df_raw = df_raw[df_raw["negara"] == negara]

    df_h = df_raw.groupby("tahun", as_index=False)["value"].sum()
    df_h.rename(columns={"tahun": "Tahun"}, inplace=True)
    df_h["Tahun"] = df_h["Tahun"].astype(str)
    df_h = df_h.sort_values("Tahun")
    df_h["Value"] = df_h["value"] / unit
    lbl = "Miliar USD" if unit > 1 else "USD"

    if metric == "yoy":
        df_h["Value"] = df_h["Value"].pct_change() * 100
        fig = px.line(df_h, x="Tahun", y="Value", markers=True, title=f"YoY (%) – HS {hs}"
                      ).update_traces(line_color=t["yellow"], marker=dict(size=8))
        fig.update_layout(**base_chart(t), yaxis_title="YoY (%)")
    else:
        fig = px.line(df_h, x="Tahun", y="Value", markers=True, title=f"Tren Nilai ({lbl}) – HS {hs}"
                      ).update_traces(line_color=t["accent"], marker=dict(size=8))
        fig.update_layout(**base_chart(t), yaxis_title=f"Nilai ({lbl})")
    return fig

# ─────────────────────────────────────────────────────────────────
#  CALLBACKS — MIRRORING
# ─────────────────────────────────────────────────────────────────
@app.callback(
    [Output("kpi-mirror","children"), Output("g-compare","figure"), Output("g-discrepancy","figure"),
     Output("tabel-mirror","data"), Output("tabel-mirror","columns"), Output("tabel-mirror","style_data_conditional"),
     Output("tabel-mirror","style_header"), Output("tabel-mirror","style_cell"), Output("status-mirror","children")],
    Input("btn-mirror","n_clicks"),
    State("dd-mitra","value"), State("dd-tahun","value"), State("radio-sumber","value"), 
    State("radio-unit-mirror","value"), State("theme-store","data"), prevent_initial_call=True,
)
def cb_mirroring(_, mitra, tahun, sumber, unit, tn):
    t   = THEMES[tn]
    bc  = base_chart(t)
    hdr = {"backgroundColor":t["card"],"color":t["text"], "fontWeight":"bold","border":f"1px solid {t['border']}"}
    cel = {"backgroundColor":t["bg"],"color":t["text"], "border":f"1px solid {t['border']}","fontFamily":FONT}
    cnd = [{"if":{"row_index":"odd"},"backgroundColor":t["card"]}]

    def _empty(msg):
        return ([], empty_fig(t), empty_fig(t), [], [], cnd, hdr, cel, msg)

    if not mitra: return _empty("⚠️ Pilih mitra dagang terlebih dahulu.")

    lbl_unit = "Juta USD" if unit > 1 else "USD"
    fmt      = "$,.1f" if unit > 1 else "$,.0f"
    tipe, bulan = get_periode_params("tahunan")

    df_bps_all = fetch_bps_db(sumber, tahun, tipe, bulan)
    if df_bps_all.empty:
        return _empty(f"⚠️ Tidak ada data BPS {tahun} di DB lokal.")

    mitra_norm = normalize_negara(mitra)
    df_bps = df_bps_all[df_bps_all["negara"] == mitra_norm].copy()
    if df_bps.empty:
        negara_ada = sorted(df_bps_all["negara"].unique().tolist())
        return _empty(html.Div([
            html.Span(f"⚠️ Tidak ada data BPS untuk '{mitra}' tahun {tahun}. ", style={"color":t["yellow"]}),
            html.Br(),
            html.Span("Negara tersedia di DB: " + ", ".join(f"'{n}'" for n in negara_ada[:20]),
                      style={"color":t["blue"],"fontSize":"11px","fontFamily":"monospace"}),
        ]))

    df_bps = df_bps.groupby("kodehs", as_index=False)["value"].sum()
    df_bps.rename(columns={"kodehs":"HS","value":"BPS_Value"}, inplace=True)
    df_bps["HS"] = df_bps["HS"].apply(clean_hs)

    df_tm, status = load_trademap(mitra, tahun, sumber)
    if status == "FILE_NOT_FOUND":
        return _empty(html.Span("❌ File 'data_trademap.xlsx' tidak ditemukan.", style={"color":t["red"]}))
    elif status == "INVALID_COLUMNS":
        return _empty(html.Span("❌ Kolom Excel salah.", style={"color":t["red"]}))
    elif status.startswith("DATA_EMPTY_TAHUN"):
        tahun_ada = status.split("|")[1] if "|" in status else "-"
        return _empty(html.Span(f"⚠️ Data Trade Map '{mitra}' tahun {tahun} tidak ada. Tersedia: {tahun_ada}", style={"color":t["yellow"]}))
    elif status == "DATA_EMPTY":
        return _empty(html.Span(f"⚠️ Mitra '{mitra}' tidak ditemukan di Excel.", style={"color":t["yellow"]}))
    elif status != "SUCCESS":
        return _empty(html.Span(f"❌ Error Excel: {status}", style={"color":t["red"]}))

    df_tm["HS"] = df_tm["HS"].apply(clean_hs)
    df = pd.merge(df_bps, df_tm, on="HS", how="outer").fillna(0)
    df["Diff_Value"] = df["Trademap_Value"] - df["BPS_Value"]
    df["HS"]         = df["HS"].apply(clean_hs)
    df["Deskripsi"]  = df["HS"].map(HS_DESC).fillna("Komoditas Lainnya")
    df = df.sort_values("BPS_Value", ascending=False)
    df[["BPS_Value","Trademap_Value","Diff_Value"]] /= unit

    tot_bps = df["BPS_Value"].sum()
    tot_tm  = df["Trademap_Value"].sum()
    selisih = tot_tm - tot_bps
    pct     = (selisih / tot_bps * 100) if tot_bps > 0 else 0
    n_match = int(((df["BPS_Value"] > 0) & (df["Trademap_Value"] > 0)).sum())
    n_bps   = int(((df["BPS_Value"] > 0) & (df["Trademap_Value"] == 0)).sum())
    n_tm    = int(((df["BPS_Value"] == 0) & (df["Trademap_Value"] > 0)).sum())

    lbl_bps = "EKSPOR IDN KE" if str(sumber) == "1" else "IMPOR IDN DARI"
    lbl_tm  = "IMPOR MITRA DARI IDN" if str(sumber) == "1" else "EKSPOR MITRA KE IDN"

    kpis = [
        html.Div(style={"background":t["card"],"border":f"1px solid {t['border']}",
                         "borderTop":f"3px solid {t['accent']}","padding":"16px","borderRadius":"8px"},
                 children=[html.Div(f"{lbl_bps} {mitra.upper()} (DB BPS)", style={"fontSize":"10px","color":t["muted"]}),
                           html.Div(f"{tot_bps:,.1f} {lbl_unit}", style={"fontSize":"20px","fontWeight":"bold","color":t["accent"]})]),
        html.Div(style={"background":t["card"],"border":f"1px solid {t['border']}",
                         "borderTop":f"3px solid {t['blue']}","padding":"16px","borderRadius":"8px"},
                 children=[html.Div(f"{lbl_tm} (Trade Map)", style={"fontSize":"10px","color":t["muted"]}),
                           html.Div(f"{tot_tm:,.1f} {lbl_unit}", style={"fontSize":"20px","fontWeight":"bold","color":t["blue"]})]),
        html.Div(style={"background":t["card"],"border":f"1px solid {t['border']}",
                         "borderTop":f"3px solid {t['yellow']}","padding":"16px","borderRadius":"8px"},
                 children=[html.Div("TOTAL ASIMETRI", style={"fontSize":"10px","color":t["muted"]}),
                           html.Div(f"{selisih:,.1f} {lbl_unit}  ({pct:+.1f}%)", style={"fontSize":"20px","fontWeight":"bold","color":t["yellow"]})]),
    ]

    top10 = df.nlargest(10, "BPS_Value").copy()
    top10["label"] = top10["HS"] + " – " + top10["Deskripsi"].str[:18]
    fig1 = go.Figure([
        go.Bar(name="DB BPS",   x=top10["label"], y=top10["BPS_Value"], marker_color=t["green"]),
        go.Bar(name="Trade Map", x=top10["label"], y=top10["Trademap_Value"], marker_color=t["red"]),
    ]).update_layout(**bc, barmode="group", height=360, xaxis=dict(tickangle=-30),
                     legend=dict(orientation="h", yanchor="bottom", y=1.02, x=1, xanchor="right"))

    df["Abs_Diff"] = df["Diff_Value"].abs()
    top5 = df.nlargest(5, "Abs_Diff").copy()
    top5["label"] = top5["HS"] + " – " + top5["Deskripsi"].str[:18]
    fig2 = go.Figure([
        go.Bar(name="DB BPS", x=top5["label"], y=top5["BPS_Value"], marker_color=t["green"],
               text=top5["BPS_Value"].apply(lambda v: f"{v:,.1f}"), textposition="outside"),
        go.Bar(name="Trade Map", x=top5["label"], y=top5["Trademap_Value"], marker_color=t["red"],
               text=top5["Trademap_Value"].apply(lambda v: f"{v:,.1f}"), textposition="outside"),
    ]).update_layout(**bc, barmode="group", height=360, xaxis=dict(tickangle=-30),
                     legend=dict(orientation="h", yanchor="bottom", y=1.02, x=1, xanchor="right"))

    cols = [
        {"name":"HS","id":"HS"}, {"name":"Deskripsi Komoditas","id":"Deskripsi"},
        {"name":f"BPS ({lbl_unit})","id":"BPS_Value", "type":"numeric","format":{"specifier":fmt}},
        {"name":f"Trade Map ({lbl_unit})","id":"Trademap_Value", "type":"numeric","format":{"specifier":fmt}},
        {"name":f"Selisih ({lbl_unit})","id":"Diff_Value", "type":"numeric","format":{"specifier":fmt}},
    ]
    cnd_tbl = cnd + [
        {"if":{"filter_query":"{Diff_Value} > 0","column_id":"Diff_Value"}, "color":t["green"]},
        {"if":{"filter_query":"{Diff_Value} < 0","column_id":"Diff_Value"}, "color":t["red"]},
    ]
    disp = ["HS","Deskripsi","BPS_Value","Trademap_Value","Diff_Value"]
    return (kpis, fig1, fig2, df[disp].to_dict("records"), cols, cnd_tbl, hdr, cel,
            html.Span(f"✅ Mirroring selesai — matched:{n_match} | BPS:{n_bps} | TM:{n_tm}", style={"color":t["green"]}))

# ─────────────────────────────────────────────────────────────────
#  CALLBACKS — DOWNLOAD
# ─────────────────────────────────────────────────────────────────
@app.callback(
    Output("dl-bps","data"), Input("btn-dl-bps","n_clicks"),
    State("store-bps","data"), State("input-negara","value"), State("input-hs","value"), prevent_initial_call=True,
)
def download_bps(_, raw, neg_f, hs_f):
    if not raw or not raw.get("data"): return None
    df = pd.DataFrame(raw["data"])
    if neg_f: df = df[df["negara"] == neg_f]
    if hs_f:  df = df[df["kodehs"] == hs_f]
    full = df.groupby(["negara","kodehs"], as_index=False)[["value","berat"]].sum()
    full["deskripsi"] = full["kodehs"].map(HS_DESC).fillna("Lainnya")
    full.rename(columns={"negara":"Negara","kodehs":"HS", "value":"Nilai_USD","berat":"Volume"}, inplace=True)
    tahun = raw.get("tahun","")
    jenis = "Ekspor" if str(raw.get("sumber")) == "1" else "Impor"
    return dcc.send_data_frame(full.to_csv, f"BPS_{jenis}_{tahun}.csv", index=False)

@app.callback(
    Output("dl-top15-csv","data"), Input("btn-dl-top15-csv","n_clicks"),
    State("store-top15","data"), prevent_initial_call=True,
)
def download_top15_csv(_, store):
    if not store or not store.get("data"): return None
    df = pd.DataFrame(store["data"])
    lbl, jenis, tahun = store.get("unit_label", "USD"), store.get("jenis", "Ekspor"), store.get("tahun", "")
    df_out = df.rename(columns={"kodehs": "Kode HS", "deskripsi": "Deskripsi Komoditas", f"value": f"Nilai ({lbl})", "berat": "Volume (kg)"})
    total = df_out[f"Nilai ({lbl})"].sum()
    df_out["Share (%)"] = (df_out[f"Nilai ({lbl})"] / total * 100).round(2)
    df_out.insert(0, "No", range(1, len(df_out) + 1))
    return dcc.send_data_frame(df_out.to_csv, f"Top15_{jenis}_{tahun}.csv", index=False)

@app.callback(
    Output("dl-top15-xlsx","data"), Input("btn-dl-top15-xlsx","n_clicks"),
    State("store-top15","data"), prevent_initial_call=True,
)
def download_top15_xlsx(_, store):
    if not store or not store.get("data"): return None
    df = pd.DataFrame(store["data"])
    xlsx_bytes = build_top15_xlsx(df, store.get("jenis", "Ekspor"), store.get("tahun", ""), store.get("unit_label", "USD"), store.get("is_ekspor", True))
    return dcc.send_bytes(lambda _: xlsx_bytes, f"Top15_{store.get('jenis', 'Ekspor')}_{store.get('tahun', '')}.xlsx")

@app.callback(
    Output("dl-mirror","data"), Input("btn-dl-mirror","n_clicks"),
    State("tabel-mirror","data"), State("dd-mitra","value"), State("dd-tahun","value"), prevent_initial_call=True,
)
def download_mirror(_, tbl_data, mitra, tahun):
    if not tbl_data: return None
    fname = f"Mirroring_{(mitra or 'mitra').replace(' ','_')}_{tahun}.csv"
    return dcc.send_data_frame(pd.DataFrame(tbl_data).to_csv, fname, index=False)

# ─────────────────────────────────────────────────────────────────
#  CALLBACKS — SEKI TAB 5
# ─────────────────────────────────────────────────────────────────
@app.callback(
    Output("seki-db-badge","children"), Input("theme-store","data"),
)
def seki_badge(tn):
    t = THEMES[tn]
    if _BOP_OK:
        return html.Div(
            f"✅  Database SEKI tersambung — data terbaru: {_BOP_LATEST} | {len(_BOP_YEARS)} tahun ({_BOP_YEARS[0]}–{_BOP_YEARS[-1]})",
            style={"fontSize":"12px","color":t["green"],"fontWeight":"bold","padding":"8px 14px","borderRadius":"6px",
                   "background":"rgba(63,185,80,0.08)","border":"1px solid rgba(63,185,80,0.3)"},
        )
    return html.Div(
        "❌  bop_indonesia.db tidak ditemukan.",
        style={"fontSize":"12px","color":t["red"],"padding":"8px 14px","borderRadius":"6px","background":"rgba(247,129,102,0.08)","border":"1px solid rgba(247,129,102,0.3)"},
    )

@app.callback(
    [Output("seki-kpi","children"), Output("seki-g-ca","figure"), Output("seki-g-wf","figure"),
     Output("seki-g-inv","figure"), Output("seki-g-cad","figure"), Output("seki-g-pct","figure"),
     Output("seki-tbl","data"), Output("seki-tbl","columns"), Output("seki-tbl","style_data_conditional"),
     Output("seki-tbl","style_header"), Output("seki-tbl","style_cell"), Output("seki-status","children")],
    Input("btn-seki","n_clicks"),
    State("seki-y1","value"), State("seki-y2","value"), State("seki-freq","value"), State("seki-unit","value"),
    State("seki-tbl-filter","value"), State("theme-store","data"), prevent_initial_call=True,
)
def cb_seki(_, y1, y2, freq, udiv, tbl_f, tn):
    t   = THEMES[tn]
    bc  = base_chart(t)
    hdr = {"backgroundColor":t["card"],"color":t["text"], "fontWeight":"bold","border":f"1px solid {t['border']}"}
    cel = {"backgroundColor":t["bg"],"color":t["text"], "border":f"1px solid {t['border']}","fontFamily":FONT,"fontSize":"11px"}
    cnd = [{"if":{"row_index":"odd"},"backgroundColor":t["card"]}]
    lbl = "Miliar USD" if udiv == 1000 else "Juta USD"

    def _fail(msg):
        ef = empty_fig(t, msg)
        return ([], ef, ef, ef, ef, ef, [], [], cnd, hdr, cel, html.Span(msg, style={"color":t["red"]}))

    if not _BOP_OK: return _fail("❌ Database bop_indonesia.db tidak ditemukan.")
    if y1 > y2: return _fail("⚠️ Tahun awal harus ≤ tahun akhir.")

    needed = sorted({1,2,17,20,23,26,29,32,35,40,41,46,47,48,54,55,56})
    df_all = bop_series(needed, y1, y2, freq)
    if df_all.empty: return _fail(f"⚠️ Tidak ada data untuk {y1}–{y2}.")

    xcol = "period" if freq == "quarterly" else "year"

    def gs(iid):
        s = df_all[df_all["item_id"] == iid].copy()
        s = s.sort_values("year" if freq == "annual" else ["year","quarter"])
        s["v"] = s["value_mn_usd"] / udiv
        return s

    def kv(iid):
        v = bop_latest_val(iid)
        return v / udiv if v is not None else None

    def kpi_card(color, label, val, sub=""):
        vstr = f"{val:,.1f} {lbl}" if val is not None else "N/A"
        return html.Div(
            style={"background":t["card"],"border":f"1px solid {t['border']}", "borderTop":f"3px solid {color}","padding":"16px","borderRadius":"8px"},
            children=[html.Div(label, style={"fontSize":"10px","color":t["muted"]}),
                      html.Div(vstr,  style={"fontSize":"18px","fontWeight":"bold","color":color}),
                      html.Div(sub,   style={"fontSize":"10px","color":t["muted"]})])

    ca_v = bop_latest_val(1)
    kpis = [
        kpi_card(t["green"] if (ca_v or 0) >= 0 else t["red"], "TRANSAKSI BERJALAN (TERBARU)", kv(1), _BOP_LATEST),
        kpi_card(t["accent"], "CADANGAN DEVISA (TERBARU)", bop_latest_val(54) / udiv if bop_latest_val(54) else None, _BOP_LATEST),
        kpi_card(t["purple"], "NERACA KESELURUHAN (TERBARU)", kv(48), "Surplus > 0"),
    ]
    ca_pdb_v = bop_latest_val(56)
    kpis.append(html.Div(
        style={"background":t["card"],"border":f"1px solid {t['border']}", "borderTop":f"3px solid {t['yellow']}","padding":"16px","borderRadius":"8px"},
        children=[html.Div("CA % PDB (TERBARU)", style={"fontSize":"10px","color":t["muted"]}),
                  html.Div(f"{ca_pdb_v:.1f}%" if ca_pdb_v is not None else "N/A", style={"fontSize":"18px","fontWeight":"bold","color":t["yellow"]}),
                  html.Div("Defisit < 0", style={"fontSize":"10px","color":t["muted"]})]))

    fig_ca = go.Figure()
    for iid, lbl_c, c in [(2,"Barang",t["green"]),(17,"Jasa",t["blue"]),(20,"Pend. Primer",t["yellow"]),(23,"Pend. Sekunder",t["orange"])]:
        s = gs(iid)
        if not s.empty: fig_ca.add_trace(go.Bar(x=s[xcol], y=s["v"], name=lbl_c, marker_color=c, opacity=0.8))
    s_ca = gs(1)
    if not s_ca.empty: fig_ca.add_trace(go.Scatter(x=s_ca[xcol], y=s_ca["v"], name="Total CA", line=dict(color=t["accent"], width=2.5), mode="lines+markers", marker=dict(size=4)))
    fig_ca.add_hline(y=0, line_dash="dash", line_color=t["muted"], line_width=1)
    fig_ca.update_layout(**bc, barmode="relative", height=340, yaxis_title=lbl, legend=dict(orientation="h", yanchor="bottom", y=1.02, bgcolor="rgba(0,0,0,0)"))

    wf_ids, wf_lbl, wf_v = [1, 26, 29, 47, 48], ["Transaksi\nBerjalan","Transaksi\nModal","Transaksi\nFinansial","Selisih\nPerhitungan","Neraca\nKeseluruhan"], []
    for iid in wf_ids:
        s = gs(iid)
        wf_v.append(float(s["v"].sum()) if not s.empty else 0)
    fig_wf = go.Figure(go.Waterfall(
        x=wf_lbl, measure=["relative","relative","relative","relative","total"], y=wf_v, connector=dict(line=dict(color=t["border"], width=1.5)),
        decreasing=dict(marker_color=t["red"]), increasing=dict(marker_color=t["green"]), totals=dict(marker_color=t["purple"]), texttemplate="%{y:,.0f}", textposition="outside",
    )).update_layout(**bc, height=340, showlegend=False, yaxis_title=lbl)

    fig_inv = go.Figure()
    for iid, lbl_i, c in [(32,"FDI",t["green"]),(35,"Portofolio",t["blue"]),(41,"Lainnya",t["orange"]),(40,"Derivatif",t["yellow"])]:
        s = gs(iid)
        if not s.empty: fig_inv.add_trace(go.Bar(x=s[xcol], y=s["v"], name=lbl_i, marker_color=c, opacity=0.85))
    s_fin = gs(29)
    if not s_fin.empty: fig_inv.add_trace(go.Scatter(x=s_fin[xcol], y=s_fin["v"], name="Total Fin.", mode="lines+markers", line=dict(color=t["accent"], width=2, dash="dot"), marker=dict(size=5)))
    fig_inv.add_hline(y=0, line_dash="dash", line_color=t["muted"], line_width=1)
    fig_inv.update_layout(**bc, barmode="relative", height=320, yaxis_title=lbl, legend=dict(orientation="h", yanchor="bottom", y=1.02, bgcolor="rgba(0,0,0,0)"))

    fig_cad = go.Figure()
    s_cad, s_ner = gs(54), gs(48)
    if not s_cad.empty: fig_cad.add_trace(go.Scatter(x=s_cad[xcol], y=s_cad["v"], name="Cadangan Devisa", fill="tozeroy", mode="lines", line=dict(color=t["teal"], width=2), fillcolor="rgba(57,208,216,0.12)"))
    if not s_ner.empty: fig_cad.add_trace(go.Bar(x=s_ner[xcol], y=s_ner["v"], name="Neraca Keseluruhan", marker_color=[t["green"] if v >= 0 else t["red"] for v in s_ner["v"]], opacity=0.8, yaxis="y2"))
    fig_cad.add_hline(y=0, line_dash="dash", line_color=t["muted"], line_width=1)
    fig_cad.update_layout(**bc, height=320, yaxis=dict(title=f"Caddev ({lbl})", gridcolor=t["grid"]), yaxis2=dict(title=f"Neraca ({lbl})", overlaying="y", side="right", gridcolor="rgba(0,0,0,0)"), legend=dict(orientation="h", yanchor="bottom", y=1.02, bgcolor="rgba(0,0,0,0)"))

    fig_pct = go.Figure()
    s_pct, s_bln = df_all[df_all["item_id"] == 56].copy().sort_values("year"), df_all[df_all["item_id"] == 55].copy().sort_values("year")
    if not s_pct.empty: fig_pct.add_trace(go.Bar(x=s_pct[xcol], y=s_pct["value_mn_usd"], name="CA % PDB", marker_color=[t["green"] if v >= 0 else t["red"] for v in s_pct["value_mn_usd"]], opacity=0.9))
    if not s_bln.empty: fig_pct.add_trace(go.Scatter(x=s_bln[xcol], y=s_bln["value_mn_usd"], name="Caddev (Bln Impor)", mode="lines+markers", line=dict(color=t["yellow"], width=2), marker=dict(size=5), yaxis="y2"))
    fig_pct.add_hline(y=0, line_dash="dash", line_color=t["muted"], line_width=1)
    fig_pct.update_layout(**bc, height=320, yaxis=dict(title="CA % PDB", gridcolor=t["grid"]), yaxis2=dict(title="Bulan Impor", overlaying="y", side="right", gridcolor="rgba(0,0,0,0)"), legend=dict(orientation="h", yanchor="bottom", y=1.02, bgcolor="rgba(0,0,0,0)"))

    if tbl_f: df_t = df_all[df_all["item_id"].isin(tbl_f)].copy()
    else: df_t = df_all.copy()
    df_t["nilai"] = df_t["value_mn_usd"] / udiv
    
    if freq == "quarterly": df_t = df_t[["period","keterangan","items_en","nilai"]].rename(columns={"period":"Periode","keterangan":"Keterangan","items_en":"English","nilai":lbl})
    else: df_t = df_t[["year","keterangan","items_en","nilai"]].rename(columns={"year":"Tahun","keterangan":"Keterangan","items_en":"English","nilai":lbl})
    
    tbl_cols = [{"name":c,"id":c, **({"type":"numeric","format":{"specifier":",.1f"}} if c == lbl else {})} for c in df_t.columns]
    cnd_tbl = cnd + [{"if":{"filter_query":f"{{{lbl}}} > 0","column_id":lbl},"color":t["green"]}, {"if":{"filter_query":f"{{{lbl}}} < 0","column_id":lbl},"color":t["red"]}]
    status_msg = html.Span(f"✅ {len(df_t):,} baris | {y1}–{y2} | {'Kuartalan' if freq=='quarterly' else 'Tahunan'} | {lbl}", style={"color":t["purple"]})
    
    return (kpis, fig_ca, fig_wf, fig_inv, fig_cad, fig_pct, df_t.to_dict("records"), tbl_cols, cnd_tbl, hdr, cel, status_msg)

@app.callback(
    Output("seki-g-cmp","figure"), Input("seki-dd-cmp","value"),
    State("seki-y1","value"), State("seki-y2","value"), State("seki-freq","value"), State("seki-unit","value"), State("theme-store","data"), prevent_initial_call=True,
)
def cb_seki_cmp(item_ids, y1, y2, freq, udiv, tn):
    t   = THEMES[tn]
    lbl = "Miliar USD" if udiv == 1000 else "Juta USD"
    if not item_ids or not _BOP_OK: return empty_fig(t, "Pilih indikator untuk membandingkan.")
    df = bop_series(item_ids, y1, y2, freq)
    if df.empty: return empty_fig(t, "Tidak ada data.")
    
    xcol = "period" if freq == "quarterly" else "year"
    pal  = [t["accent"],t["green"],t["red"],t["yellow"],t["purple"],t["teal"],t["orange"],t["blue"]]
    fig  = go.Figure()
    
    for i, iid in enumerate(item_ids):
        s = df[df["item_id"] == iid].copy().sort_values("year" if freq == "annual" else ["year","quarter"])
        if s.empty: continue
        name = BOP_MAIN_ITEMS.get(iid, s["keterangan"].iloc[0])
        fig.add_trace(go.Scatter(x=s[xcol], y=s["value_mn_usd"] / udiv, name=name, mode="lines+markers", line=dict(color=pal[i % len(pal)], width=2), marker=dict(size=5)))
        
    fig.add_hline(y=0, line_dash="dot", line_color=t["muted"], line_width=1)
    fig.update_layout(**base_chart(t), height=320, yaxis_title=lbl, legend=dict(orientation="h", yanchor="bottom", y=1.02, bgcolor="rgba(0,0,0,0)"))
    return fig

@app.callback(
    Output("dl-seki","data"), Input("btn-dl-seki","n_clicks"),
    State("seki-tbl","data"), State("seki-y1","value"), State("seki-y2","value"), prevent_initial_call=True,
)
def dl_seki(_, data, y1, y2):
    if not data: return None
    return dcc.send_data_frame(pd.DataFrame(data).to_csv, f"SEKI_BI_{y1}_{y2}.csv", index=False)

if __name__ == "__main__":
    app.run(debug=False, host="0.0.0.0", port=PORT)
